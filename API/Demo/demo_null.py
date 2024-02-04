# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
# @Time   : 2024/01/27 09:45
# @Author : 新贵大人
描述:
"""
import base64
import json
import os
import re
import time
import jsonpath
import requests
import openpyxl
from datetime import datetime, timedelta
from qcloud_cos import CosConfig, CosS3Client
from pathlib import Path

class PushStudentOrder:
    requests.packages.urllib3.disable_warnings()
    def __init__(self):
        '''
        :param environment: 环境uat or preprod
        :param num: 需要上传的学员订单条数
        '''
        path=os.getcwd()
        file_path = path+r"\test.xlsx"
        self.path = file_path
        self.environment = self.input_environment()
        self.num = self.input_num()
        self.token = self.get_crm_login_token()
        self.front_sign_data = None

    def get_base64(self, passwd):
        b_pwd = base64.b64encode(passwd.encode()).decode('utf-8')
        return b_pwd

    def input_environment(self):
        while True:
            environment = input("请输入：(1)-->测试环境    (2)-->预发布环境")
            try:
                environment = int(environment)
            except ValueError:
                print("输入错误，请输入数字")
                continue
            if environment == 1:
                return "uat"
            elif environment == 2:
                return "preprod"
            else:
                print("请输入一个数字(1)/(2)获取正确的环境")
                continue

    def input_num(self):
        while True:
            num = input(f"请输入{self.environment}环境需要导入的订单条数:")
            try:
                num = int(num)
            except ValueError:
                print("输入错误，请输入数字")
                continue
            if 0 < num <= 100:
                return num
            else:
                print("请输入大于0，小于100的数字")
                continue

    def handle_api_response(self, response):
        try:
            resp_json = response.json()
            response.raise_for_status()
            return resp_json
        except requests.exceptions.RequestException as e:
            return {"error": f"API请求错误: {e}"}

    # def get_path(self,file_name):
    #     file_local_path_name = os.path.join(os.path.dirname(os.path.realpath(__file__)), f"{file_name}.xlsx")
    #     return file_local_path_name

    # def get_path(self,file_name):
    #     SRC_PATH = Path.absolute(Path(__file__)).parent  # 获取临时目录路径对象
    #     file_path = str(SRC_PATH / f"{file_name}.xlsx")  # 拼接获得文件绝对路径字符串
    #     return file_path

    # def get_path1(self):
    #     SRC_PATH = Path.absolute(Path(__file__)).parent  # 获取临时目录路径对象
    #     file_path = str(SRC_PATH / "test.xlsx")  # 拼接获得文件绝对路径字符串
    #     aaa = window.setWindowIcon(QIcon(file_path))
    #     print("aaa",aaa)
    #     print(file_path)

    def get_crm_login_token(self, username=18707699952, passw='klzz1234@@'):
        url = f'https://{self.environment}-auth-new.vipthink.cn/iam-sso/v2/auth/admin/token'
        data = {
            "account": username,
            "password": self.get_base64(passw),
            "loginType": "acc_pwd"
        }
        token = requests.post(url, json=data,verify=False).json().get('data', {}).get('token', '')
        return token

    def get_student_phone(self):
        phone_list = []
        try:
            for i in range(self.num):
                now = time.time()
                new_phone = '1' + str(now).replace('.', '')[-10:]
                url = f'https://{self.environment}-gw.vipthink.cn/api/member/v3/back/ol-user/getUserInfoByMobile'
                payload = {"mobile": f"{new_phone}"}
                resp = self.handle_api_response(requests.post(url, json=payload, headers={"authorization": self.token},verify=False))
                if resp and resp.get('code') == 0 and not resp.get('data'):
                    if phone_list != []:
                        # 判断列表中有没有相同的手机号码
                        if new_phone in [i for i in phone_list]:
                            if int(new_phone[-1:]) != 9:
                                # \d表示匹配一个数字字符，$表示匹配字符串的结尾位置
                                phone_list.append(re.sub(r'\d$', str(int(new_phone[-1:]) + 1), new_phone))
                            else:
                                phone_list.append(re.sub(r'\d$', str(int(new_phone[-1:]) - 1), new_phone))
                        else:
                            phone_list.append(new_phone)
                    else:
                        phone_list.append(new_phone)
        except TypeError as error:
            print('API request exception, please check\n', 'error:', error)
        return phone_list

    def operation_table(self):
        data_list = self.get_student_phone()
        file_path = self.path
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook["Sheet1"]
        formatted_datetime1 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_datetime2 = datetime.now().strftime("%Y%m%d%H%M%S")
        # 清空表格的数据
        if sheet.max_row > 2:
            # 行号是从1开始计数的，而不是从0开始，所以这里是3，
            for row in range(3, sheet.max_row + 1):
                sheet.delete_rows(row)
        if self.environment == "uat":
            for index, value in enumerate(data_list):
                sheet.cell(row=3 + index, column=1, value=f'XG{formatted_datetime2}{value}')  # 第三方订单号
                sheet.cell(row=3 + index, column=2, value='0')  # 收款渠道
                sheet.cell(row=3 + index, column=4, value="86")  # 手机区号
                sheet.cell(row=3 + index, column=5, value=value)  # 手机号
                sheet.cell(row=3 + index, column=8, value='31825515')  # 套餐skuid
                sheet.cell(row=3 + index, column=9, value='0')  # 订单支付金额
                sheet.cell(row=3 + index, column=10, value=formatted_datetime1)  # 支付时间
                sheet.cell(row=3 + index, column=11, value="free")  # 支付方式
                sheet.cell(row=3 + index, column=12, value="0")  # 渠道id
                sheet.cell(row=3 + index, column=13, value="1999")  # 获得原因
                sheet.cell(row=3 + index, column=15, value="0")  # 是否需要地址
        elif self.environment == "preprod":
            for index, value in enumerate(data_list):
                sheet.cell(row=3 + index, column=1, value=f'XG{formatted_datetime2}{value}')  # 第三方订单号
                sheet.cell(row=3 + index, column=2, value='0')  # 收款渠道
                sheet.cell(row=3 + index, column=4, value="86")  # 手机区号
                sheet.cell(row=3 + index, column=5, value=value)  # 手机号
                sheet.cell(row=3 + index, column=8, value='20528990')  # 套餐skuid
                sheet.cell(row=3 + index, column=9, value='0.02')  # 订单支付金额
                sheet.cell(row=3 + index, column=10, value=formatted_datetime1)  # 支付时间
                sheet.cell(row=3 + index, column=11, value="free")  # 支付方式
                sheet.cell(row=3 + index, column=12, value="470659")  # 渠道id
                sheet.cell(row=3 + index, column=13, value="1999")  # 获得原因
                sheet.cell(row=3 + index, column=15, value="0")  # 是否需要地址
        workbook.save(file_path)
        workbook.close()

    def get_front_sign(self):
        if self.front_sign_data is not None:
            return self.front_sign_data

        url = f'https://{self.environment}-attch-api.vipthink.cn/v1/attach/getSign'
        data = {
            "name": "test.xlsx",
            "dir": "uploads/images",
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "ext": "xlsx",
            "size": 16323,
            "driver": "tencent_oss",
            "code": "platform-trade"
        }
        resp = self.handle_api_response(requests.post(url, json=data, headers={"authorization": self.token},verify=False))
        if resp and 'data' in resp:
            # 如果 'data' 不存在，设置为空字典
            data_dict = resp.get('data', {})
            self.front_sign_data = {
                "id": data_dict.get('id', ''),
                "ossDomain": data_dict.get('ossDomain', ''),
                "path": data_dict.get('path', ''),
                "ossToken": data_dict.get('ossToken', '')
            }
        return self.front_sign_data

    def import_order(self):
        self.update_file_to_cos()
        front_sign_data = self.get_front_sign()
        if front_sign_data:
            url2 = f'https://{self.environment}-attch-api.vipthink.cn/v1/attach/notify'
            url3 = f'https://{self.environment}-gw.vipthink.cn/api/trade_order/v1/admin/orderImport/urlImport'
            data2 = {"id": front_sign_data['id'], "status": 1}
            self.handle_api_response(requests.post(url2, json=data2, headers={"authorization": self.token},verify=False))
            data3 = {
                "fileUrl": f"{front_sign_data['ossDomain']}{front_sign_data['path']}",
                "fileName": "test.xlsx",
                "operatorId": 667508,
                "operatorName": "谭新贵",
                "importType": 1
            }
            resp = requests.post(url3, json=data3, headers={"authorization": self.token},verify=False).json()
            resp_hasFail = jsonpath.jsonpath(resp, '$..hasFail')
            if resp_hasFail:  # hasFail 为True代表文件上传失败了
                fail_details = resp['data']['importFailDtls']
                for detail in fail_details:
                    print(f'文件上传失败，手机号：{detail["mobile"]}, 错误：{detail["failReason"]}')
            else:
                print("上传文件成功")

    def update_file_to_cos(self, Bucket='test-1253622427', region='ap-guangzhou',
                           domain='.cos.accelerate.myqcloud.com'):
        # 修改要上传表格的数据
        self.operation_table()
        front_sign_data = self.get_front_sign()
        if front_sign_data and isinstance(front_sign_data['ossToken'], str):
            try:
                dic = json.loads(front_sign_data['ossToken'])
                secret_id = dic['credentials']['tmpSecretId']
                secret_key = dic['credentials']['tmpSecretKey']
                token = dic['credentials']['sessionToken']

                config = CosConfig(Region=region, SecretId=secret_id, SecretKey=secret_key, Token=token,
                                   Domain=Bucket + domain)
                client = CosS3Client(config)
                with open(self.path, 'rb') as fp:
                    response = client.put_object(
                        Bucket='test-1253622427',
                        Body=fp,
                        Key=front_sign_data['path'],
                        StorageClass='STANDARD',
                        ContentType='text/html; charset=utf-8'
                    )
                    return response['ETag']
            except Exception as e:
                return e

    @property
    def get_batch(self):
        url = f'https://{self.environment}-gw.vipthink.cn/api/trade_order/v1/admin/orderImport/importRecord/list'
        data = {"pageNo": 1, "pageSize": 20}
        resp = self.handle_api_response(requests.post(url, json=data, headers={"authorization": self.token},verify=False))
        data = resp.get('data', {})[0]
        # 将 createTime 字符串转换为 datetime 对象
        create_time = datetime.strptime(data.get('createTime', ''), "%Y-%m-%d %H:%M:%S")
        current_time = datetime.now()
        one_minute = timedelta(minutes=20)
        # 获取两分钟内创建的第一个新订单
        if current_time - create_time < one_minute and data.get('fileName', '') == "test.xlsx":
            return data.get('batchNum', '')

    def get_order_recordDtlId(self) -> list:
        url = f'https://{self.environment}-gw.vipthink.cn/api/trade_order/v1/admin/orderImport/importRecord/dtlList'
        data = {"pageNo": 1, "pageSize": 20, "batchNum": f"{self.get_batch}", "status": "WAIT_APPROVAL"}
        resp = self.handle_api_response(requests.post(url, json=data, headers={"authorization": self.token},verify=False))
        data = resp.get('data', {})
        # 提取所有recordDtlId
        record_dtl_ids = [item['recordDtlId'] for item in data]
        return record_dtl_ids

    def last_auditing_order(self):
        url = f'https://{self.environment}-gw.vipthink.cn/api/trade_order/v1/admin/orderImport/importRecord/auditDtls'
        for i in self.get_order_recordDtlId():
            data = {
                "recordDtlIds": [f"{i}"],
                "batchNum": f"{self.get_batch}",
                "auditStatus": "APPROVAL_SUCCESS",
                "remarks": "测试数据"
            }
            resp = self.handle_api_response(requests.post(url, json=data, headers={"authorization": self.token},verify=False))
            if resp.get('msg') == "success" and resp.get('message') == "OK":
                print('订单号：', f"{jsonpath.jsonpath(resp, '$..orderNumber')}""审核成功")
            else:
                print('订单号：', f"{jsonpath.jsonpath(resp, '$..orderNumber')}""审核失败")


if __name__ == '__main__':
    try:
        push_student_order = PushStudentOrder()
        push_student_order.import_order()
        push_student_order.last_auditing_order()
        time.sleep(1000)
    except Exception as e:
        print(f"异常：{e}")
        name = "test.xlsx"
        dangqianmulu = os.getcwd()
        goujianwanlujin = os.path.join(dangqianmulu,name)
        print("当前目录",dangqianmulu)
        print("当前寻找的文件路径",goujianwanlujin)
        time.sleep(1000)



